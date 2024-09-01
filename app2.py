from flask import Flask, render_template, request
import openpyxl

app = Flask(__name__)



@app.route('/')
def welcome():
    return render_template("1.html")
    

@app.route("/submit", methods = ["POST","GET"])
def calcutaion():

    height = float(request.form['height'])
    weight = float(request.form['weight'])
    gender = request.form['gender']
    age  = float(request.form['age'])

    HC_range = None

    

    agedata = openpyxl.load_workbook("static//circum.xlsx")
    age_data = agedata['Sheet1']

    wtdata = openpyxl.load_workbook("static//std_dev.xlsx")
    wt_data = wtdata['Sheet1']
    
        
    # writing like this doesnt work - return hhh,ppp
    # have to make a list out of it - but lets just take the variables and make a web oage of it 

    for row in range(1,81):
        if age_data.cell(row,1).value == age and gender =="m":
            weight_range = [ age_data.cell(row-1,2).value, age_data.cell(row,2).value,age_data.cell(row+1,2).value ]
            wt_ratio= round(weight / float(age_data.cell(row,2).value) , 2)

            height_range = [age_data.cell(row-1,3).value , age_data.cell(row,3).value ,age_data.cell(row+1,3).value ]
            ht_ratio= round(height/ float(age_data.cell(row,3).value) , 2)

            
            HC_range = [ age_data.cell(row-1,4).value, age_data.cell(row,4).value,age_data.cell(row+1,4).value ]
        
        if age_data.cell(row,1).value == age  and gender == "f":
            weight_range =[ age_data.cell(row-1,5).value, age_data.cell(row,5).value,age_data.cell(row+1,5).value ] 
            wt_ratio= round(weight / float(age_data.cell(row,5).value), 2)
    
            
            height_range = [age_data.cell(row-1,6).value, age_data.cell(row,6).value,age_data.cell(row+1,6).value ]
            ht_ratio= round(height/ float(age_data.cell(row,6).value) , 2)
            
            #HC 
            
            HC_range = [age_data.cell(row-1,7).value, age_data.cell(row,7).value,age_data.cell(row+1,7).value ]


    wfh = None
    if age <= 4 :
        adj_ht = int(round(height))
        for row2 in range(2,78) :
            
            if adj_ht == wt_data.cell(row2,6).value and gender == "m":
                wfh = round(weight / wt_data.cell(row2 , 5).value,2)
                
            if adj_ht == wt_data.cell(row2,6).value and gender == "f":
                wfh = round(weight / wt_data.cell(row2 , 7).value,2)


    bmi = None
    if age > 4 :
        h_m = ( height /100) **2
        bmi = round(weight /h_m , 2)
            

    return render_template("result.html", weight_range = weight_range ,wt_ratio = wt_ratio , height_range = height_range, ht_ratio=ht_ratio, HC_range =HC_range, wfh = wfh ,bmi = bmi)





    
        
    # return "checking if something else returned as well" - no not returned 

    '''val = float(request.form['height'])
    sentence = "the height is " + str(val)
    return sentence'''
    


if __name__ == '__main__' :
    app.run(host ="0.0.0.0", debug=True)
